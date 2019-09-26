import openpyxl as xls
# new xls
# wb=xls.Workbook()
# sheet = wb.active
# sheet.title='수집 데이터'
# sheet['B2'] = 'b2'
# sheet.cell(row=3,column=3).value='3,3'
# sheet.append([1,2,3,4,5])
# wb.save('excelfile.xlsx')
wb = xls.load_workbook('excelfile.xlsx')
sheet1 = wb.active
sheet1.title="이름변경"
sheet1.append(range(10))
wb.save('excelfile.xlsx')